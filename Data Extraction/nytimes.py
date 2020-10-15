import openpyxl
import requests
import datetime
from time import sleep
from urllib.parse import quote

query_list = ['Netease Inc Ads', 'Nvidia Corp', 'Nxp Semiconductors', "O'Reilly Automotive", 'Paychex Inc', 'Paccar Inc', 'Pepsico Inc','Paypal Holdings','Qualcomm Inc','Regeneron Pharmaceuticals','Ross Stores Inc','Starbucks Corp','Seattle Genetics Inc','Sirius XM Holdings','Synopsys Inc','Splunk Inc','Skyworks Solutions','Trip.com Group Ltd','T-Mobile US','Tesla Inc','Take-Two Interacti','Texas Instruments','United Continental Holdings','Ulta Beauty Inc','Verisk Analytics Inc','Verisign Inc','Vertex Pharmaceutic','Walgreens Boots Alliance','Workday Inc','Western Digital Cp']
api_key = ## YOUR API KEY HERE ##

for query in query_list:
    #Configuration
    begin_date = datetime.datetime.strptime("2014-01-01", "%Y-%m-%d")
    end_date = begin_date + datetime.timedelta(days=10)
    # query = "nasdaq"
    stop_date = datetime.datetime.strptime("2020-02-23", "%Y-%m-%d")


    wb = openpyxl.Workbook()
    sheet = wb.active
    cell = sheet.cell(row = 1, column = 1)
    cell.value = 'pub_date'
    cell = sheet.cell(row = 1, column = 2)
    cell.value = 'abstract'
    cell = sheet.cell(row = 1, column = 3)
    cell.value = 'lead_paragraph'
    cell = sheet.cell(row = 1, column = 4)
    cell.value = 'snippet'

    try:
        stop_condition = True
        cell_row = 2
        while (stop_condition):
            if (end_date >= stop_date):
                stop_condition = False
                end_date = stop_date
            # print(begin_date.strftime("%Y%m%d"), " - " , end_date.strftime("%Y%m%d"))
            #api call
            url = "https://api.nytimes.com/svc/search/v2/articlesearch.json?q=" + quote(query) + "&begin_date=" + begin_date.strftime("%Y%m%d") + "&end_date=" + end_date.strftime("%Y%m%d") + "&api-key=" + api_key + "&page=0"
            payload = {}
            headers= {}
            response = requests.request("GET", url, headers=headers, data = payload)
            data = response.json()

            for i in range(len(data['response']['docs'])):
                cell = sheet.cell(row = cell_row, column = 1)
                cell.value = data['response']['docs'][i]['pub_date']
                cell = sheet.cell(row = cell_row, column = 2)
                cell.value = data['response']['docs'][i]['abstract']
                cell = sheet.cell(row = cell_row, column = 3)
                cell.value = data['response']['docs'][i]['lead_paragraph']
                cell = sheet.cell(row = cell_row, column = 4)
                cell.value = data['response']['docs'][i]['snippet']
                cell_row += 1
            #pause for 7 seconds to not hit nytimes api limit
            sleep(7)

            if (data['response']['meta']['hits'] > 10):
                for page in range(1, data['response']['meta']['hits']//10 + 1):
                    url = "https://api.nytimes.com/svc/search/v2/articlesearch.json?q=" + query + "&begin_date=" + begin_date.strftime("%Y%m%d") + "&end_date=" + end_date.strftime("%Y%m%d") + "&api-key=" + api_key + "&page=" + str(page)
                    new_response = requests.request("GET", url, headers=headers, data = payload)
                    new_data = new_response.json()
                    for i in range(len(new_data['response']['docs'])):
                        cell = sheet.cell(row = cell_row, column = 1)
                        cell.value = new_data['response']['docs'][i]['pub_date']
                        cell = sheet.cell(row = cell_row, column = 2)
                        cell.value = new_data['response']['docs'][i]['abstract']
                        cell = sheet.cell(row = cell_row, column = 3)
                        cell.value = new_data['response']['docs'][i]['lead_paragraph']
                        cell = sheet.cell(row = cell_row, column = 4)
                        cell.value = new_data['response']['docs'][i]['snippet']
                        cell_row += 1
                    #pause for 7 seconds to not hit nytimes api limit
                    sleep(7)

            begin_date = end_date + datetime.timedelta(days=1)
            end_date += datetime.timedelta(days=10)
            

        wb.save(query + ".xlsx")
    except:
        cell = sheet.cell(row = 1, column = 1)
        cell.value = 'Did not extract fully'
        wb.save(query + ".xlsx")

